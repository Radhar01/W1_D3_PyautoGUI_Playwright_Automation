# playwright_assign.py
# Assignment 2 - Playwright WhatsApp Web Bot
# Beginner version (written in a simple, step-by-step way).
#
# What it does:
# 1. Opens WhatsApp Web (scan the QR code the first time).
# 2. Reads contacts from contacts.xlsx (columns: Name, Phone, Message).
# 3. For each contact: searches them, sends a personalized message,
#    waits, and takes a screenshot.
# 4. Reads the last 3 messages from the chat.
# 5. Saves a report as JSON and as Excel.
#
# NOTE: Only message people who agreed to hear from you.

import time
import random
import json
from datetime import datetime

import openpyxl
from playwright.sync_api import sync_playwright

# ---------------- settings ----------------
today = datetime.now().strftime("%Y-%m-%d")
json_report_file = "whatsapp_report_" + today + ".json"
excel_report_file = "whatsapp_report_" + today + ".xlsx"

# WhatsApp Web selectors (these are the boxes on the page).
# WhatsApp changes its internal markup fairly often, so instead of one
# hard-coded selector we try a short list of known variants and use
# whichever one shows up first. If WhatsApp changes its website again and
# all of these stop working, open DevTools (F12) on web.whatsapp.com,
# click the box you need, and add the new selector to the matching list.
# NOTE: newer WhatsApp Web builds replaced the old contenteditable-div
# search box with a plain <input aria-label="Search or start a new chat">,
# so both the old and new variants are listed here.
LOGGED_IN_SELECTORS = [
    '#pane-side',
    '#side',
]
SEARCH_SELECTORS = [
    'input[aria-label="Search or start a new chat"]',
    'div[contenteditable="true"][aria-label="Search or start a new chat"]',
    'div[contenteditable="true"][data-tab="3"]',
    'div[contenteditable="true"][aria-label="Search input textbox"]',
    'div[contenteditable="true"][data-tab="2"][aria-label="Search input textbox"]',
    '[aria-label="Search input textbox"]',
]
MESSAGE_SELECTORS = [
    'div[contenteditable="true"][data-tab="10"]',
    'div[contenteditable="true"][aria-label="Type a message"]',
    'footer div[contenteditable="true"][role="textbox"]',
    'footer div[contenteditable="true"]',
]


# a small helper so the bot pauses like a human (2 to 5 seconds)
def human_pause():
    time.sleep(random.randint(2, 5))


# tries each selector in turn, repeatedly, until one appears or time runs out
def wait_for_any(page, selectors, timeout_ms):
    deadline = time.time() + (timeout_ms / 1000)
    last_error = None
    while time.time() < deadline:
        for selector in selectors:
            try:
                element = page.wait_for_selector(selector, timeout=3000)
                if element:
                    return element
            except Exception as e:
                last_error = e
    raise Exception(
        "None of these selectors showed up in time: "
        + str(selectors)
        + ". WhatsApp Web may have changed its layout again - "
        + "open DevTools and check the current attributes. "
        + "Last error: " + str(last_error)
    )


# ---------------- step 1: read the contacts from Excel ----------------
print("Reading contacts.xlsx ...")
wb = openpyxl.load_workbook("contacts.xlsx")
sheet = wb.active

contacts = []
# row 1 is the header, so we start reading from row 2
for row in sheet.iter_rows(min_row=2, values_only=True):
    name = row[0]
    phone = row[1]
    message = row[2]
    if name is None:          # skip empty rows
        continue
    contacts.append({"name": name, "phone": phone, "message": message})

print("Found", len(contacts), "contacts")

# this list will hold the result for every contact (for the report)
report = []


# ---------------- step 2: open the browser ----------------
with sync_playwright() as p:
    # user_data_dir keeps you logged in, so you only scan the QR code once
    browser = p.chromium.launch_persistent_context(
        "whatsapp_login",   # a folder where the login is saved
        headless=False      # we want to SEE the browser
    )
    page = browser.new_page()
    page.goto("https://web.whatsapp.com")

    print("If needed, scan the QR code with your phone...")
    # wait up to 2 minutes for the chat list to appear (means we are logged in)
    wait_for_any(page, LOGGED_IN_SELECTORS, timeout_ms=120000)
    print("Logged in!")

    # ---------------- step 3: go through each contact ----------------
    for contact in contacts:
        name = contact["name"]
        phone = contact["phone"]
        template = contact["message"]

        # make the personalized message (replace {name} with the real name)
        if template:
            message = template.replace("{name}", name)
        else:
            message = "Hi " + name + ", this is an automated message."

        print("\nContact:", name)

        try:
            # --- search the contact by name (you could also use the phone) ---
            search = wait_for_any(page, SEARCH_SELECTORS, timeout_ms=15000)
            search.click()
            search.fill("")            # clear the box
            search.type(str(name))
            page.wait_for_timeout(2000)  # wait for results to show

            # click the chat that has this name. Scoped to the side panel and
            # narrowed to the first match, since a plain "span[title=name]"
            # can match more than one element and Playwright refuses to
            # click when a selector is ambiguous.
            side_panel = page.locator("#pane-side")
            side_panel.get_by_text(str(name), exact=True).first.click()
            human_pause()

            # --- type and send the message ---
            box = wait_for_any(page, MESSAGE_SELECTORS, timeout_ms=15000)
            box.click()
            box.type(message)
            page.keyboard.press("Enter")
            print("Message sent.")
            human_pause()

            # --- take a screenshot of the chat ---
            shot_name = "sent_" + str(name).replace(" ", "_") + ".png"
            page.screenshot(path=shot_name)

            # --- read the last 3 messages in the chat ---
            bubbles = page.query_selector_all(
                'span[data-testid="selectable-text"], span.selectable-text'
            )
            last_texts = []
            for bubble in bubbles[-3:]:
                last_texts.append(bubble.inner_text())

            # save the result for the report
            report.append({
                "name": name,
                "phone": phone,
                "message": message,
                "status": "sent",
                "screenshot": shot_name,
                "last_3_messages": last_texts
            })

        except Exception as e:
            # if anything goes wrong (like contact not found), don't crash
            print("Could not message", name, "-", e)
            report.append({
                "name": name,
                "phone": phone,
                "message": message,
                "status": "failed",
                "screenshot": "",
                "last_3_messages": []
            })

        human_pause()

    print("\nAll contacts done. Closing browser soon...")
    page.wait_for_timeout(3000)
    browser.close()


# ---------------- step 4: save the JSON report ----------------
print("Saving JSON report...")
with open(json_report_file, "w", encoding="utf-8") as f:
    json.dump(report, f, indent=4, ensure_ascii=False)

# ---------------- step 5: save the Excel report ----------------
print("Saving Excel report...")
out_wb = openpyxl.Workbook()
out_sheet = out_wb.active
out_sheet.append(["Name", "Phone", "Message", "Status", "Screenshot"])
for r in report:
    out_sheet.append([r["name"], r["phone"], r["message"], r["status"], r["screenshot"]])
out_wb.save(excel_report_file)

print("Done!")
print("Saved:", json_report_file)
print("Saved:", excel_report_file)
