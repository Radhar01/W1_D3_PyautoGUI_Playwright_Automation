# playwright_assign_manual.py
# Assignment 2 - Playwright WhatsApp Web Bot (manual/flat version)
# Same single-file logic as playwright_assign.py, written top-to-bottom
# with no helper functions.
#
# What it does (per the assignment brief):
# 1. Logs in to WhatsApp Web (scan the QR code manually on the first run).
# 2. Reads contacts from contacts.xlsx (columns: Name, Phone, Message).
# 3. For each contact: searches them by name or phone number, sends a
#    personalized message, waits and confirms it was actually sent
#    (checks for the sent/delivered tick), then screenshots it.
# 4. Smart data extraction: reads back the last 3 messages in the chat.
# 5. Saves a report as both JSON and Excel, with a status per contact
#    (sent / contact not found / failed) so a bad contact never crashes
#    the whole run.
#
# NOTE: Only message people who agreed to hear from you.

import time
import random
import json
from datetime import datetime

import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

# ---------------- settings ----------------
today = datetime.now().strftime("%Y-%m-%d")
json_report_file = "whatsapp_report_" + today + ".json"
excel_report_file = "whatsapp_report_" + today + ".xlsx"

# WhatsApp Web selectors (these are the boxes on the page).
# If WhatsApp changes its website, you may need to update these.
# Each has a fallback (comma = "or" in a CSS selector) since WhatsApp
# occasionally swaps its internal data-tab numbering. Newer builds replaced
# the old contenteditable-div search box with a plain
# <input aria-label="Search or start a new chat">, so that variant is
# listed first since it's what current WhatsApp Web actually renders.
LOGGED_IN_MARKER = '#pane-side, #side'
SEARCH_BOX = 'input[aria-label="Search or start a new chat"], div[contenteditable="true"][data-tab="3"], div[aria-label="Search input textbox"]'
MESSAGE_BOX = 'div[contenteditable="true"][data-tab="10"], div[aria-label="Type a message"]'
SENT_TICK = 'span[data-icon="msg-check"], span[data-icon="msg-dblcheck"], span[data-icon="msg-time"]'


# ---------------- step 1: read the contacts from Excel ----------------
print("Reading contacts.xlsx ...")
wb = openpyxl.load_workbook("contacts.xlsx")
sheet = wb.active

contacts = []
# row 1 is the header, so we start reading from row 2
for row in sheet.iter_rows(min_row=2, values_only=True):
    name = row[0]
    phone = row[1]
    message = row[2]
    if name is None:          # skip empty rows
        continue
    contacts.append({"name": name, "phone": phone, "message": message})

print("Found", len(contacts), "contacts")

# this list will hold the result for every contact (for the report)
report = []


# ---------------- step 2: open the browser ----------------
with sync_playwright() as p:
    # user_data_dir keeps you logged in, so you only scan the QR code once
    browser = p.chromium.launch_persistent_context(
        "whatsapp_login",   # a folder where the login is saved
        headless=False      # we want to SEE the browser
    )
    page = browser.new_page()
    page.goto("https://web.whatsapp.com")

    print("If needed, scan the QR code with your phone...")
    # wait up to 2 minutes for the chat list to appear (means we are logged in)
    page.wait_for_selector(LOGGED_IN_MARKER, timeout=120000)
    print("Logged in!")

    # ---------------- step 3: go through each contact ----------------
    for contact in contacts:
        name = contact["name"]
        phone = contact["phone"]
        template = contact["message"]

        # make the personalized message (replace {name} with the real name)
        if template:
            message = template.replace("{name}", name)
        else:
            message = "Hi " + name + ", this is an automated message."

        print("\nContact:", name)

        try:
            # --- search the contact by name, or by phone number if given ---
            search = page.wait_for_selector(SEARCH_BOX)
            search.click()
            search.fill("")                    # clear the box
            search_query = str(phone) if phone else str(name)
            search.type(search_query)
            # let results load - a real wait, and also our human-like pause
            page.wait_for_timeout(random.randint(2, 5) * 1000)

            # click the chat that matches. WhatsApp shows the saved name in
            # the title if the number is a known contact, otherwise it shows
            # the phone number instead - so try the name first, then the
            # number, so search-by-name and search-by-number both work.
            # Scoped to the side panel and narrowed to the first match,
            # since an unscoped "span[title=...]" can match more than one
            # element and Playwright refuses to click an ambiguous selector.
            side_panel = page.locator("#pane-side")
            try:
                side_panel.get_by_text(str(name), exact=True).first.click(timeout=5000)
            except PlaywrightTimeoutError:
                side_panel.get_by_text(str(phone), exact=True).first.click(timeout=5000)

            time.sleep(random.randint(2, 5))   # pause like a human

            # --- type and send the message ---
            box = page.wait_for_selector(MESSAGE_BOX)
            box.click()
            box.type(message)
            page.keyboard.press("Enter")

            # confirm the message was actually sent (a tick icon shows up
            # on the bubble once WhatsApp has queued/sent it)
            page.wait_for_selector(SENT_TICK, timeout=15000)
            print("Message sent.")
            time.sleep(random.randint(2, 5))   # pause like a human

            # --- take a screenshot of the sent message ---
            shot_name = "sent_" + str(name).replace(" ", "_") + ".png"
            page.screenshot(path=shot_name)

            # --- smart extraction: read the last 3 messages in the chat ---
            bubbles = page.query_selector_all(
                'span[data-testid="selectable-text"], span.selectable-text'
            )
            last_texts = []
            for bubble in bubbles[-3:]:
                last_texts.append(bubble.inner_text())

            # save the result for the report
            report.append({
                "name": name,
                "phone": phone,
                "message": message,
                "status": "sent",
                "screenshot": shot_name,
                "last_3_messages": last_texts
            })

        except PlaywrightTimeoutError:
            # the search/click/send never found what it was waiting for -
            # most likely the contact just isn't on WhatsApp / not found
            print("Could not message", name, "- contact not found")
            report.append({
                "name": name,
                "phone": phone,
                "message": message,
                "status": "contact not found",
                "screenshot": "",
                "last_3_messages": []
            })

        except Exception as e:
            # anything else unexpected - don't let one bad contact crash the run
            print("Could not message", name, "-", e)
            report.append({
                "name": name,
                "phone": phone,
                "message": message,
                "status": "failed",
                "screenshot": "",
                "last_3_messages": []
            })

        time.sleep(random.randint(2, 5))   # pause before the next contact

    print("\nAll contacts done. Closing browser soon...")
    page.wait_for_timeout(3000)
    browser.close()


# ---------------- step 4: save the JSON report ----------------
print("Saving JSON report...")
with open(json_report_file, "w", encoding="utf-8") as f:
    json.dump(report, f, indent=4, ensure_ascii=False)

# ---------------- step 5: save the Excel report ----------------
print("Saving Excel report...")
out_wb = openpyxl.Workbook()
out_sheet = out_wb.active
out_sheet.append(["Name", "Phone", "Message", "Status", "Screenshot"])
for r in report:
    out_sheet.append([r["name"], r["phone"], r["message"], r["status"], r["screenshot"]])
out_wb.save(excel_report_file)

print("Done!")
print("Saved:", json_report_file)
print("Saved:", excel_report_file)
